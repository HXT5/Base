'''
excel
openpyxl:.xlsx读写操作
安装pip install openpyxl
操作except的三个对象：
工作薄（Workbook）
表单（sheet）
单元格（cell）
'''
import openpyxl
import os
file=os.path.join(os.path.dirname(os.path.abspath(__file__)),'data.xlsx')
#1、加载excel数据文件
wb=openpyxl.load_workbook(file)
#2、根据表单名称选择表单，wb['表单名称']
sheet=wb['login']
#3、在表单中，获取单元格数据
    #3、1 单元格对象：sh.cell(row,colum) #下标从1开始
    #3、2 .value获取单元格的值
cell=sheet.cell(1,1).value
print(cell)
#4、得到当前表单中总行数和总列表
print(sheet.max_row)
print(sheet.max_column)

#获取excel的title
#5读取整行数据sheet.rows，是生成器转为list
list_title=[]
for item in list(sheet.rows)[0]:
    list_title.append(item.value)
data=[]
#获取除了title的单元格数据
for item in list(sheet.rows)[1:]:
    list_list = []
    for cell in item:
       list_list.append(cell.value)
    res=dict(zip(list_title,list_list))
    data.append(res)
print(data)










